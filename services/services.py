from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side


def create_excel(tuple_list: list, output_filename: str, sheet_name="Sheet1"):

    filepath = os.path.join("services/tmp", output_filename)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [
        "Время",
        "МО",
        "Головная заявка",
        "Дочерняя заявка",
        "Краткое описание проблемы",
        "Решение",
    ]

    # Записываем данные из списка словарей в Excel
    if tuple_list:
        ws.append(headers)  # Записываем заголовки

        for row in tuple_list:
            columns = []
            for col in row:
                if type(col) is datetime:
                    columns.append(str(col))
                else:
                    columns.append(col)
            ws.append(columns)

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_style.border = border_style

    cell_style = NamedStyle(name="cell")
    cell_style.alignment = Alignment(horizontal="center", vertical="center")
    cell_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(filepath)
    return filepath
